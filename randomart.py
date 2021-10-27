import numpy as np
from numpy.random import default_rng
from openpyxl import Workbook
from openpyxl import load_workbook

n = int(input("how many articles do you want per time?"))  # number of articles

#load search results and create empty output file
befarticles = load_workbook(filename="Resultatenlijst voor_brexit.xlsx")
aftarticles = load_workbook(filename="Resultatenlijst after_brexit.xlsx")
selected = Workbook()

#get active sheets
befarticlesheet = befarticles.active
afarticlesheet = aftarticles.active
selectedsheet = selected.active

#make some rows and columns
selectedsheet.insert_rows(2*n)
selectedsheet.insert_cols(4)

#random number set between 1 and 1000 (arange set accordingly)
rng = default_rng()
articleset = rng.choice(np.arange(1,1001), size=n, replace=False)
articleset2 = rng.choice(np.arange(1,1001), size=n, replace=False)

#define proxy url
proxylink = "https://advance-lexis-com.proxy-ub.rug.nl"

#we loop through and add each line to the output
for i in range(1,n+1):
  for j in range(1,5):
    #get value from original sheet
    c1 = befarticlesheet.cell(row=articleset[i-1], column=j)
    c2 = afarticlesheet.cell(row=articleset2[i-1], column=j)

    #put in output
    selectedsheet.cell(row=i,column=j).value = c1.value
    selectedsheet.cell(row=i+n,column=j).value = c2.value

    #we also want to get the links
    if j==1:
      hyperlink1 = c1.hyperlink
      hyperlink2 = c2.hyperlink

      #make the links proxy-friendly
      hyperlink1.target = proxylink + hyperlink1.target[25:]
      hyperlink2.target = proxylink + hyperlink2.target[25:]

      #put in new sheet
      selectedsheet.cell(row=i,column=j).hyperlink = hyperlink1
      selectedsheet.cell(row=i+n,column=j).hyperlink = hyperlink2

#save the result
selected.save(filename="telegraph_bonus_extra.xlsx")

